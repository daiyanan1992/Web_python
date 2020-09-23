from openpyxl import load_workbook

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name


    def get_data(self):
        # 打开excel
        wb = load_workbook(filename=self.file_name)
        # 定位sheet
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(1, sheet.max_row + 1):
            sub_data = {}
            sub_data['method'] = sheet.cell(i + 1, 1).value
            sub_data['url'] = sheet.cell(i + 1, 2).value
            sub_data['data'] = sheet.cell(i + 1, 3).value
            test_data.append(sub_data)
        return test_data

if __name__ == '__main__':
    res = DoExcel(r'D:\Web_python2\Stu_01\Homework\daiyn.xlsx','python1').get_data()
    print(res)





