from openpyxl import load_workbook



class DoExcel:
    # def __init__(self,file_name,sheet_name,result):
    #     self.file_name=file_name
    #     self.sheet_name=sheet_name
    #     self.result=result

    def get_data(self,file_name,sheet_name):

        #1先打开excel
        wb = load_workbook(file_name)
        #2定位到sheet页
        sheet = wb[sheet_name]
        #3查询数据
        test_data = []
        for i in range(1, sheet.max_row):#遍历最大的行数
            sub_data = {}
            sub_data['num'] = sheet.cell(i + 1,1).value
            sub_data['case_name'] = sheet.cell(i + 1,2).value
            sub_data['method'] = sheet.cell(i + 1, 3).value
            sub_data['url'] = sheet.cell(i + 1, 4).value
            sub_data['data'] = sheet.cell(i + 1, 5).value
            sub_data['expected'] = sheet.cell(i + 1,6).value
            test_data.append(sub_data)
        return test_data

    def save_data(self,file_name,sheet_name,i,result):
        # 1先打开excel
        wb = load_workbook(file_name)
        # 2定位到sheet页
        sheet = wb[sheet_name]
        # for i in range(1,sheet.max_row):
        sheet.cell(i + 1,7).value = result
        wb.save(file_name)



if __name__ == '__main__':
    res = DoExcel().get_data(r'D:\Web_python2\Python_jehu\test_data\test_case925.xlsx','juhe')
    print(res)
    # DoExcel(r'D:\Web_python2\Python_jehu\test_data\test_case925.xlsx', 'juhe','daiyananis shuaige').save_data()